import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


CSV_PATH = r"D:\football_analysis\data\football_match_collection.csv"
XLSX_PATH = r"D:\football_analysis\data\football_match_collection_review.xlsx"


def main() -> None:
    with open(CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    section_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(vertical="center", horizontal="center")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sections = [
        (
            "基础信息",
            [
                "match_id",
                "league_or_issue",
                "home_team",
                "away_team",
                "match_time",
                "source_match_url",
                "shuju_url",
                "ouzhi_url",
                "touzhu_url",
            ],
        ),
        (
            "维度一：基础实力",
            [
                "elo_home",
                "elo_away",
                "recent_form_home",
                "recent_form_away",
                "home_away_form",
            ],
        ),
        (
            "维度二：近期动态",
            [
                "head_to_head_summary",
                "injury_or_lineup_notes",
                "motivation_or_schedule_notes",
            ],
        ),
        (
            "维度三：市场数据",
            [
                "european_odds_movement_summary",
                "betting_heat_summary",
            ],
        ),
        ("来源与备注", ["media_source_links", "collected_sources", "remarks"]),
    ]

    summary = wb.create_sheet("总览")
    summary["A1"] = "检阅版总览"
    summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary["A1"].fill = header_fill
    summary.merge_cells("A1:D1")

    headers = ["工作表", "match_id", "比赛", "比赛时间"]
    for idx, header in enumerate(headers, start=1):
        cell = summary.cell(2, idx, header)
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = border
        cell.alignment = center

    used_titles = set()
    for row_idx, row in enumerate(rows, start=3):
        sheet_title = f"{row['home_team']}vs{row['away_team']}"[:31]
        if sheet_title in used_titles:
            suffix = 2
            base = sheet_title[:28]
            while f"{base}_{suffix}" in used_titles:
                suffix += 1
            sheet_title = f"{base}_{suffix}"
        used_titles.add(sheet_title)

        summary_values = [
            sheet_title,
            row["match_id"],
            f"{row['home_team']} vs {row['away_team']}",
            row["match_time"],
        ]
        for col_idx, value in enumerate(summary_values, start=1):
            cell = summary.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = wrap

        ws = wb.create_sheet(sheet_title)
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

        ws["A1"] = "比赛检阅表"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = header_fill
        ws["A1"].alignment = center
        ws.merge_cells("A1:B1")

        ws["A2"] = "比赛"
        ws["B2"] = f"{row['home_team']} vs {row['away_team']}"
        ws["A2"].font = Font(bold=True)
        ws["A2"].fill = section_fill
        ws["A2"].border = border
        ws["A2"].alignment = center
        ws["B2"].border = border
        ws["B2"].alignment = wrap

        current_row = 3
        for section_name, fields in sections:
            ws.cell(current_row, 1, section_name)
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).fill = section_fill
            ws.cell(current_row, 1).border = border
            ws.cell(current_row, 1).alignment = wrap
            ws.cell(current_row, 2, "")
            ws.cell(current_row, 2).fill = section_fill
            ws.cell(current_row, 2).border = border
            current_row += 1

            for field in fields:
                ws.cell(current_row, 1, field)
                ws.cell(current_row, 2, row.get(field, ""))
                ws.cell(current_row, 1).font = Font(bold=True)
                ws.cell(current_row, 1).border = border
                ws.cell(current_row, 1).alignment = wrap
                ws.cell(current_row, 2).border = border
                ws.cell(current_row, 2).alignment = wrap
                current_row += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 120

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 32
    summary.column_dimensions["D"].width = 20

    wb.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
